#!/usr/bin/env python3
"""
extract_all_2025.py

Extracts IFRS 17 financial data from 2025 annual reports for:
  - Achmea BV
  - a.s.r. Nederland N.V.
  - NN Group N.V.
  - Athora Netherlands N.V.

Output: All_Insurers_2025_Databook.xlsx  (one tab per insurer)

Usage:  python extract_all_2025.py
Needs:  ANTHROPIC_API_KEY in .env or environment
Deps:   pdfplumber, anthropic, openpyxl
"""

import json
import os
import re
import time
import concurrent.futures
from pathlib import Path
from threading import Semaphore

import pdfplumber
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

OUT_PATH   = "All_Insurers_2025_Databook.xlsx"
MODEL      = "claude-opus-4-6"
MAX_TOKENS = 4096

# Load API key from .env / .Renviron — checks project dir then home dir
def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text().splitlines():
        line = line.strip()
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

for _candidate in [Path(".env"), Path(".Renviron"),
                   Path.home() / ".env", Path.home() / ".Renviron"]:
    _load_env_file(_candidate)

API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
if not API_KEY:
    raise SystemExit("ANTHROPIC_API_KEY not set. Add it to .env or .Renviron.")

client = anthropic.Anthropic(api_key=API_KEY)

# Cap concurrent API calls at 3 to avoid rate-limit backoffs while still
# running all 4 insurers in parallel for PDF extraction and assembly.
_api_sem = Semaphore(3)

# ---------------------------------------------------------------------------
# Page maps
# Each dict entry is a list of 1-based page numbers matching the PDF.
# Comments document what is on each page.
# ---------------------------------------------------------------------------

PAGE_MAPS = {
    "achmea": {
        "pdf":               "Annual report Achmea BV.pdf",
        "short_name":        "Achmea",
        "balance_sheet":     [198],
        "income_statement":  [199],
        "portfolio_overview": [253],           # Note 7 summary (GMM/PAA/VFA by segment)
        "nonlife_movements": [257, 259],        # p257=total NL 2025, p259=GMM NL 2025
        "health_movements":  [263],             # p263=total Health 2025
        "life_movements":    [267, 269, 270],
        "csm_rollforward":   [259, 269, 270],   # Non-Life GMM + Life GMM/VFA (2025 only)
        "ra_rollforward":    [257, 259, 263, 267, 269],
        "loss_component":    [257, 259, 263, 267, 269],
        "csm_maturity":      [254],
        "insurance_svc_result": [296, 297],
        "net_financial_result": [298, 299],
        "discount_rates":    [275, 276, 277],   # p275=UFR text, p276=rate table, p277=confidence
        "solvency":          [34, 35],
        "investments_note":  [248, 249, 250, 251, 252],
        "gross_premium":     [30, 33, 39],
        # Discount curve: table on p276 has three EUR curves; GMM Euro -> liquid, Life/VFA Euro -> illiquid
        # Dutch comma decimals; return midpoint of range; UFR=2.3%, CoC=4.5%
        "disc_liquid_label":    "GMM Euro",
        "disc_illiquid_label":  "Life Netherlands GMM and VFA",
    },

    "asr": {
        "pdf":               "2025-annual-report-asr.pdf",
        "short_name":        "a.s.r.",
        "balance_sheet":     [262],
        "income_statement":  [263],
        "portfolio_overview": [315, 316],       # Note 7.5.13: Insurance contract liabilities
        "nonlife_movements": [316, 317, 319],   # Non-life: total + GMM component tables 2025
        "life_movements":    [321, 322, 323],   # Life: total + GMM component tables 2025
        "csm_rollforward":   [319, 323],        # Non-life GMM (319) + Life GMM/VFA (323) 2025
        "ra_rollforward":    [316, 319, 321, 323],
        "loss_component":    [316, 317, 321, 322],
        "csm_maturity":      [332, 339],        # p332=GMM maturity (7.5.13.6); p339=VFA maturity (7.5.14.3)
        "insurance_svc_result": [347, 348],     # p347=7.6.1 insurance contract revenue (CSM/RA release)
        "net_financial_result": [350],          # p350=section 7.6.7 investment and finance result
        "discount_rates":    [329, 330],        # p329=assumptions (UFR=3.20%, FSP=20y); p330=rate table
        "solvency":          [406],             # EOF=12,618, SCR=5,743, ratio=220%
        "investments_note":  [306, 307, 308],
        "gross_premium":     [286, 287, 288, 295, 296, 297],
        # p330 table rows: "0% (min)" = liquid, "100% (max)" = illiquid
        "disc_liquid_label":    "0% (min)",
        "disc_illiquid_label":  "100% (max)",
    },

    "nn": {
        "pdf":               "nn-group-annual-report-2025.pdf",
        "short_name":        "NN Group",
        "balance_sheet":     [161, 162],
        "income_statement":  [163],
        "portfolio_overview": [192, 221],       # p192=model breakdown, p221=segment breakdown
        "nonlife_movements": [200, 201],        # PAA movements 2025
        "life_movements":    [193, 194],        # GMM/VFA movements 2025
        "csm_rollforward":   [193, 195, 196],   # GMM/VFA with CSM disaggregation
        "ra_rollforward":    [193, 197, 200],
        "loss_component":    [197, 198, 200, 201],
        "csm_maturity":      [195, 196],
        "insurance_svc_result": [163, 164],
        "net_financial_result": [163, 164],
        "discount_rates":    [191, 192],        # Discount curve + confidence levels
        "solvency":          [37, 38, 45],
        "investments_note":  [173, 174, 175, 176],
        "gross_premium":     [163, 221],        # p221: GWP Life=8,787 NL=4,469 Total=13,256
        # p191 columns: "General Model" (liquid), "Variable Fee Approach" (illiquid)
        # LTFR=3.20% (=ufr), LLP=30y (=fsp_years), CoC=4%
        "disc_liquid_label":    "General Model",
        "disc_illiquid_label":  "Variable Fee Approach",
    },

    "athora": {
        "pdf":               "annual-report-athora-netherlands-nv-2025.pdf",
        "short_name":        "Athora NL",
        "balance_sheet":     [80],
        "income_statement":  [81],              # revenue=2,110, service result=262, PBT=-237
        # 100% Life (GMM+VFA). No Non-life, no Health, no PAA.
        # p119: GMM=29,255, VFA=14,132, Total=43,387
        "portfolio_overview": [80, 119],
        "nonlife_movements": [120],             # kept for schema compat
        "life_movements":    [120, 122],        # p120=LRC table 2025; p122=measurement component 2025
        "csm_rollforward":   [122],
        "ra_rollforward":    [122],
        "loss_component":    [120],
        "csm_maturity":      [131, 132],        # Athora does not publish CSM-release-by-time-bucket
        "insurance_svc_result": [81, 120, 122],
        "net_financial_result": [81],
        "discount_rates":    [129, 130],        # p129=rate table (Liquid/Illiquid exact headers)
        "solvency":          [177, 178],        # p177: SII ratio=197%, EOF=3,532, SCR=1,790
        "investments_note":  [99, 100],         # bonds=16,313, mortgages=3,282, total=66,502
        "gross_premium":     [14, 81],
        # p129 exact column headers: "Liquid" and "Illiquid"
        "disc_liquid_label":    "Liquid",
        "disc_illiquid_label":  "Illiquid",
    },
}

# ---------------------------------------------------------------------------
# PDF extraction helpers
# ---------------------------------------------------------------------------

def _table_to_markdown(table: list[list]) -> str:
    """
    Convert a pdfplumber table (list of rows, each row a list of cell strings)
    to a pipe-delimited Markdown table.
    Skips entirely-None rows. Fills None cells with empty string.
    """
    rows = []
    for row in table:
        cleaned = [str(c).strip() if c is not None else "" for c in row]
        if any(cleaned):
            rows.append(cleaned)
    if not rows:
        return ""

    # Determine column widths for alignment
    n_cols = max(len(r) for r in rows)
    rows = [r + [""] * (n_cols - len(r)) for r in rows]  # pad short rows

    def fmt_row(cells):
        return "| " + " | ".join(cells) + " |"

    lines = [fmt_row(rows[0])]
    lines.append("|" + "|".join(["---"] * n_cols) + "|")
    for row in rows[1:]:
        lines.append(fmt_row(row))
    return "\n".join(lines)


def extract_page(pdf_pages: dict, page_num: int) -> str:
    """
    Extract content from a single (1-based) page.
    Tables are converted to Markdown; remaining text is appended below.
    Strips navigation chrome common in Dutch insurer reports.
    """
    page = pdf_pages[page_num]
    parts = []

    # --- Extract tables first -------------------------------------------------
    tables = page.extract_tables(
        table_settings={
            "vertical_strategy":   "lines_strict",
            "horizontal_strategy": "lines_strict",
            "snap_tolerance":      4,
            "join_tolerance":      4,
            "edge_min_length":     10,
        }
    )
    # Fall back to text-based detection if no lattice lines found
    if not tables:
        tables = page.extract_tables(
            table_settings={
                "vertical_strategy":   "text",
                "horizontal_strategy": "text",
                "snap_tolerance":      6,
                "join_tolerance":      6,
                "edge_min_length":     20,
            }
        )

    table_bboxes = []
    for tbl in tables:
        md = _table_to_markdown(tbl)
        if md:
            parts.append(md)
            # Record approximate bbox to avoid re-extracting table area as prose
            # (pdfplumber does not return bbox per table in extract_tables; we
            #  use find_tables() for that when needed, but simple dedup suffices here)

    # --- Extract remaining text -----------------------------------------------
    # Use find_tables to get bboxes so we can exclude those regions from text
    found = page.find_tables(
        table_settings={
            "vertical_strategy":   "lines_strict",
            "horizontal_strategy": "lines_strict",
            "snap_tolerance":      4,
            "join_tolerance":      4,
            "edge_min_length":     10,
        }
    )
    if not found:
        found = page.find_tables(
            table_settings={
                "vertical_strategy":   "text",
                "horizontal_strategy": "text",
                "snap_tolerance":      6,
                "join_tolerance":      6,
                "edge_min_length":     20,
            }
        )

    # Crop page to exclude table bounding boxes before text extraction
    page_for_text = page
    for t in found:
        try:
            page_for_text = page_for_text.outside_bbox(t.bbox)
        except Exception:
            pass

    prose = page_for_text.extract_text(x_tolerance=3, y_tolerance=3) or ""
    prose = _clean_prose(prose)
    if prose.strip():
        parts.append(prose)

    return "\n\n".join(parts)


def _clean_prose(text: str) -> str:
    """Remove navigation chrome that appears on every page of Dutch insurer reports."""
    lines = text.splitlines()
    cleaned = []
    for line in lines:
        # Header band: "Annual accounts ... Other information ... Appendix"
        if re.search(r"Annual accounts.{1,60}Other information.{1,60}Appendix", line):
            continue
        # Footer: "Annual Report 20XX | NNN"
        if re.search(r"Annual Report 20\d{2}\s*\|\s*\d{1,4}\s*$", line):
            continue
        # Collapse 3+ spaces to 2 (reduces tokens, preserves column separators)
        line = re.sub(r" {3,}", "  ", line)
        if line.strip():
            cleaned.append(line)
    return "\n".join(cleaned)


def open_pdf(pdf_path: str) -> dict:
    """
    Open a PDF and return a dict mapping 1-based page numbers to pdfplumber page objects.
    Also logs page count for parity with R script output.
    """
    pdf = pdfplumber.open(pdf_path)
    pages = {i + 1: p for i, p in enumerate(pdf.pages)}
    print(f"  {len(pages)} pages loaded.")
    return pages


def extract_pages(pdf_pages: dict, page_nums: list[int]) -> str:
    """Extract and concatenate a list of pages, separated by PAGE BREAK markers."""
    parts = []
    for p in page_nums:
        if p not in pdf_pages:
            continue
        content = extract_page(pdf_pages, p)
        if content.strip():
            parts.append(content)
    return "\n\n---PAGE BREAK---\n\n".join(parts)


# ---------------------------------------------------------------------------
# Claude API wrapper
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = (
    "You are a specialist in IFRS 17 insurance financial statement analysis. "
    "Extract ONLY the numeric values explicitly stated in the provided text. "
    "Return ONLY valid JSON — no markdown fences, no explanation. "
    "Use null for any value not found. All monetary values in EUR millions. "
    "Percentages as decimals (e.g. 4.5% = 0.045). "
    "Return integers where possible (no decimal for whole numbers). "
    "Numbers may use Dutch/European comma decimal notation (2,17 means 2.17) — convert to dot decimals. "
    "Min-max ranges (e.g. 2.17-2.41) should be returned as their midpoint unless instructed otherwise."
)

# Regex to extract first flat JSON object — immune to prose before/after fence.
# [^{}]* matches only the first self-contained flat JSON block (no nested {}).
_JSON_RE = re.compile(r"\{[^{}]*\}", re.DOTALL)


def call_claude(prompt_text: str, section_name: str, insurer_name: str) -> dict | None:
    """Call Claude and parse the first JSON object from the response."""
    label = f"{insurer_name} / {section_name}"
    print(f"  Calling Claude for: {label} ...")

    last_err = None
    for attempt in range(3):
        try:
            with _api_sem:
                response = client.messages.create(
                    model=MODEL,
                    max_tokens=MAX_TOKENS,
                    system=[
                        {
                            "type": "text",
                            "text": SYSTEM_PROMPT,
                            "cache_control": {"type": "ephemeral"},
                        }
                    ],
                    messages=[{"role": "user", "content": prompt_text}],
                )
                raw = response.content[0].text
            break
        except anthropic.RateLimitError as e:
            wait = 15 * (attempt + 1)
            print(f"  Rate limit hit for {label}, waiting {wait}s ...")
            time.sleep(wait)
            last_err = e
        except anthropic.APIStatusError as e:
            if e.status_code in (500, 502, 503, 529):
                wait = 15 * (attempt + 1)
                print(f"  API error {e.status_code} for {label}, waiting {wait}s ...")
                time.sleep(wait)
                last_err = e
            else:
                raise
    else:
        print(f"  WARNING: all retries failed for {label}: {last_err}")
        return None

    # Extract the first flat JSON object from the response
    m = _JSON_RE.search(raw.strip())
    if not m:
        print(f"  WARNING: no JSON found for {label}. Raw: {raw[:200]}")
        return None
    try:
        return json.loads(m.group())
    except json.JSONDecodeError as e:
        print(f"  WARNING: JSON parse failed for {label}: {e}\n  Raw: {raw[:300]}")
        return None


# ---------------------------------------------------------------------------
# Extraction function — all Claude calls for one insurer
# ---------------------------------------------------------------------------

def extract_insurer(pm: dict) -> dict:
    insurer = pm["short_name"]
    print(f"\n{'=' * 60}")
    print(f"Loading PDF: {pm['pdf']}")
    pdf_pages = open_pdf(pm["pdf"])

    res = {}

    # ---- S1: Portfolio overview ------------------------------------------------
    print(f"\n[S1] Portfolio overview ({insurer})...")
    s1_text = extract_pages(pdf_pages, pm["portfolio_overview"])
    res["s1"] = call_claude(
        f"From this {insurer} 2025 annual report extract (IFRS 17 insurance contract overview), "
        "extract 31 December 2025 net balance sheet values (EUR millions, net = liabilities minus assets). "
        "Look for a summary table showing insurance contract liabilities by measurement model. "
        "The table may have rows for Life insurance contracts and Non-life insurance contracts, "
        "with columns for General Model (GMM), Variable Fee Approach (VFA), Premium Allocation Approach (PAA) and Total. "
        "Use the net total row (liabilities minus assets) for each segment.\n\n"
        'Return JSON:\n'
        '{"life_gmm": <Life GMM net>, "life_vfa": <Life VFA net>, "life_paa": <Life PAA net>,'
        ' "life_total": <Life total net>,'
        ' "nonlife_gmm": <Non-Life GMM net>, "nonlife_paa": <Non-Life PAA net>, "nonlife_total": <Non-Life total net>,'
        ' "health_gmm": <Health GMM net or null>, "health_paa": <Health PAA net or null>, "health_total": <Health total net or null>,'
        ' "total_gmm": <total GMM>, "total_vfa": <total VFA>, "total_paa": <total PAA>,'
        ' "total_direct": <grand total direct insurance contracts net>,'
        ' "reins_total": <total outward reinsurance contracts held net or null>}\n\n'
        f"Text:\n{s1_text}",
        "portfolio_overview", insurer
    )

    # ---- S2a: Insurance service result ----------------------------------------
    print(f"\n[S2a] Insurance service result ({insurer})...")
    s2a_pages = list(dict.fromkeys(pm["income_statement"] + pm["insurance_svc_result"]))
    s2a_text = extract_pages(pdf_pages, s2a_pages)
    res["s2a"] = call_claude(
        f"From this {insurer} 2025 annual report extract (income statement and insurance service result note), "
        "extract 2025 values (EUR millions).\n\n"
        "Return JSON:\n"
        '{"net_insurance_result": <total insurance service result 2025>,'
        ' "csm_release_total": <CSM recognised for services provided>,'
        ' "ra_release_total": <Change in risk adjustment for non-financial risk released>,'
        ' "paa_insurance_revenue": <PAA insurance revenue total>,'
        ' "paa_incurred_claims": <PAA incurred claims total>,'
        ' "gmm_incurred_claims": <GMM incurred claims/benefits total>,'
        ' "net_reinsurance_result": <Net result from reinsurance contracts>,'
        ' "net_financial_result": <Net financial result from insurance activities>,'
        ' "profit_before_tax": <Profit before tax>,'
        ' "insurance_revenue_total": <Insurance revenue total>}\n\n'
        f"Text:\n{s2a_text}",
        "insurance_service_result", insurer
    )

    # ---- S2b: Discount rates ---------------------------------------------------
    print(f"\n[S2b] Discount rates ({insurer})...")
    s2d_text = extract_pages(pdf_pages, pm["discount_rates"])
    res["s2d"] = call_claude(
        f"From this {insurer} 2025 annual report extract (discount curve section), "
        "extract discount rate parameters at 31 December 2025 for EUR insurance contracts.\n\n"
        "RULES:\n"
        "- Convert Dutch comma decimals (2,17 -> 2.17).\n"
        "- Convert percentages to decimals (3.20% -> 0.032).\n"
        "- For min-max ranges (e.g. 2,17-2,41) return the midpoint.\n"
        f"- '{pm['disc_liquid_label']}' maps to liquid_* keys.\n"
        f"- '{pm['disc_illiquid_label']}' maps to illiquid_* keys.\n"
        "- UFR / Ultimate Forward Rate / LTFR (Long-Term Forward Rate): convert to decimal and return as ufr.\n"
        "- FSP / First Smoothing Point / Last Liquid Point (LLP): return as integer years and return as fsp_years.\n"
        "- Cost of capital rate: convert to decimal.\n"
        "- Confidence levels: look for % figures linked to Non-Life, Life, Health.\n\n"
        "Return JSON:\n"
        '{"liquid_1y": <number or null>, "liquid_5y": <number or null>, "liquid_10y": <number or null>,'
        ' "liquid_15y": <number or null>, "liquid_20y": <number or null>,'
        ' "liquid_30y": <number or null>, "liquid_40y": <number or null>, "liquid_50y": <number or null>,'
        ' "illiquid_1y": <number or null>, "illiquid_5y": <number or null>, "illiquid_10y": <number or null>,'
        ' "illiquid_15y": <number or null>, "illiquid_20y": <number or null>,'
        ' "illiquid_30y": <number or null>, "illiquid_40y": <number or null>, "illiquid_50y": <number or null>,'
        ' "ufr": <decimal or null>, "fsp_years": <integer or null>,'
        ' "cost_of_capital_rate": <decimal or null>,'
        ' "confidence_nonlife": <decimal or null>, "confidence_life": <decimal or null>,'
        ' "confidence_health": <decimal or null>}\n\n'
        f"Text:\n{s2d_text}",
        "discount_rates", insurer
    )

    # ---- S3: LRC deep-dive (CSM + RA + Loss component) — single merged call ---
    print(f"\n[S3] LRC deep-dive ({insurer})...")
    s3_page_nums = sorted(set(
        pm["csm_rollforward"] + pm["ra_rollforward"] + pm["loss_component"]
    ))
    s3_text = extract_pages(pdf_pages, s3_page_nums)
    res["s3"] = call_claude(
        f"From this {insurer} 2025 annual report extract (LRC/LIC movement tables), "
        "extract ALL THREE of the following for 2025 (EUR millions). "
        "Extract 2025 figures only — ignore 2024 comparative columns/tables.\n\n"
        "(A) CSM ROLLFORWARD: CSM may appear as dedicated columns or embedded in GMM tables.\n"
        "  csm_opening_total = opening balance; csm_new_business = contracts initially recognised;\n"
        "  csm_future_service_changes = changes in estimates adjusting CSM;\n"
        "  csm_finance_result = finance result on CSM; csm_release = CSM for services provided (negative);\n"
        "  csm_other = FX/acquisitions/other; csm_closing_total = closing balance.\n"
        "  Also split by segment where available: csm_*_nonlife and csm_*_life.\n\n"
        "(B) RISK ADJUSTMENT (RA) ROLLFORWARD: RA as standalone table or column.\n"
        "  ra_opening_total, ra_new_business, ra_future_service_adj_csm, ra_future_service_no_csm,\n"
        "  ra_past_service, ra_finance_result, ra_release, ra_other, ra_closing_total.\n"
        "  Segment splits: ra_opening_nonlife, ra_closing_nonlife, ra_opening_life, ra_closing_life.\n\n"
        "(C) LOSS COMPONENT: column in LRC movement tables.\n"
        "  lc_opening_total, lc_losses_recognised, lc_systematic_alloc, lc_future_service_changes,\n"
        "  lc_finance, lc_other, lc_closing_total.\n"
        "  Segment splits: lc_opening_nonlife, lc_closing_nonlife, lc_opening_life, lc_closing_life,\n"
        "  lc_opening_health, lc_closing_health.\n\n"
        "Return ONE JSON with all fields (use null for any value not found):\n"
        '{"csm_opening_total": <n>, "csm_new_business": <n>, "csm_future_service_changes": <n>,'
        ' "csm_finance_result": <n>, "csm_release": <n>, "csm_other": <n>, "csm_closing_total": <n>,'
        ' "csm_opening_nonlife": <n>, "csm_closing_nonlife": <n>,'
        ' "csm_release_nonlife": <n>, "csm_new_business_nonlife": <n>,'
        ' "csm_opening_life": <n>, "csm_closing_life": <n>,'
        ' "csm_release_life": <n>, "csm_new_business_life": <n>,'
        ' "ra_opening_total": <n>, "ra_new_business": <n>,'
        ' "ra_future_service_adj_csm": <n>, "ra_future_service_no_csm": <n>,'
        ' "ra_past_service": <n>, "ra_finance_result": <n>,'
        ' "ra_release": <n>, "ra_other": <n>, "ra_closing_total": <n>,'
        ' "ra_opening_nonlife": <n>, "ra_closing_nonlife": <n>,'
        ' "ra_opening_life": <n>, "ra_closing_life": <n>,'
        ' "lc_opening_total": <n>, "lc_losses_recognised": <n>,'
        ' "lc_systematic_alloc": <n>, "lc_future_service_changes": <n>,'
        ' "lc_finance": <n>, "lc_other": <n>, "lc_closing_total": <n>,'
        ' "lc_opening_nonlife": <n>, "lc_closing_nonlife": <n>,'
        ' "lc_opening_life": <n>, "lc_closing_life": <n>,'
        ' "lc_opening_health": <n>, "lc_closing_health": <n>}\n\n'
        f"Text:\n{s3_text}",
        "lrc_deepdive", insurer
    )

    # ---- S3d: CSM maturity -----------------------------------------------------
    print(f"\n[S3d] CSM maturity ({insurer})...")
    s3d_text = extract_pages(pdf_pages, pm["csm_maturity"])
    res["s3d"] = call_claude(
        f"From this {insurer} 2025 annual report extract, find the CSM maturity or coverage period breakdown "
        "showing how much CSM is expected to be released per time bucket (EUR millions, 2025 year-end). "
        "If there are separate GMM and VFA tables, sum them together for each bucket.\n\n"
        "Return JSON:\n"
        '{"csm_maturity_lt1y": <0-1 year>, "csm_maturity_1to5y": <1-5 years>,'
        ' "csm_maturity_5to10y": <5-10 years or null>, "csm_maturity_gt10y": <over 10 years or null>,'
        ' "csm_maturity_total": <total>}\n\n'
        f"Text:\n{s3d_text}",
        "csm_maturity", insurer
    )

    # ---- S4: Solvency ----------------------------------------------------------
    print(f"\n[S4] Solvency ({insurer})...")
    s4_text = extract_pages(pdf_pages, pm["solvency"])
    res["s4"] = call_claude(
        f"From this {insurer} 2025 annual report extract, find the Solvency II ratio and capital at 31 December 2025.\n\n"
        "Return JSON:\n"
        '{"solvency2_ratio": <decimal e.g. 1.82 for 182%>, "solvency2_target": <decimal or null>,'
        ' "scr": <EUR millions or null>, "eligible_own_funds": <EUR millions or null>,'
        ' "capital_generated": <EUR millions or null>}\n\n'
        f"Text:\n{s4_text}",
        "solvency", insurer
    )

    # ---- S5: Investments -------------------------------------------------------
    print(f"\n[S5] Investments ({insurer})...")
    s5_text = extract_pages(pdf_pages, pm["investments_note"])
    res["s5"] = call_claude(
        f"From this {insurer} 2025 annual report extract, extract the investment asset mix at 31 December 2025 "
        "(EUR millions). Classify by asset type. Separate own-risk (insurance) from policyholder / unit-linked if possible.\n\n"
        "Return JSON:\n"
        '{"equities": <number or null>, "govt_bonds": <number or null>, "corporate_bonds": <number or null>,'
        ' "mortgages": <number or null>, "other_fixed_income": <number or null>,'
        ' "derivatives_net": <number or null>, "other_investments": <number or null>,'
        ' "total_investments": <total>, "fvoci": <number or null>, "fvtpl": <number or null>,'
        ' "amortised_cost": <number or null>}\n\n'
        f"Text:\n{s5_text}",
        "investments", insurer
    )

    # ---- S6: Gross written premium ---------------------------------------------
    print(f"\n[S6] Gross written premium ({insurer})...")
    s6_text = extract_pages(pdf_pages, pm["gross_premium"])
    res["s6"] = call_claude(
        f"From this {insurer} 2025 annual report extract, find gross written premium (GWP) or insurance revenue "
        "split by Life, Non-Life, Health/Disability and total for 2025 (EUR millions). "
        "If GWP is not available, use insurance revenue totals by segment as the best proxy. "
        "Look for segment tables, notes with premium income, or consolidated income statement lines "
        "that break out Life / Non-Life / Health / Pensions / International.\n\n"
        "Return JSON:\n"
        '{"gwp_life": <number or null>, "gwp_health": <number or null>,'
        ' "gwp_nonlife": <number or null>, "gwp_pensions": <number or null>,'
        ' "gwp_intl": <number or null>, "gwp_total": <number>}\n\n'
        f"Text:\n{s6_text}",
        "gross_premium", insurer
    )

    return res


# ---------------------------------------------------------------------------
# Excel assembly
# ---------------------------------------------------------------------------

HEADER_FONT   = Font(bold=True, size=11)
TITLE_FONT    = Font(bold=True, size=14)
SUBHEAD_FONT  = Font(bold=True, size=10)
SECTION_FILL  = PatternFill("solid", fgColor="D9E1F2")   # light blue
SUBTOT_FILL   = PatternFill("solid", fgColor="E2EFDA")   # light green


def write_insurer_tab(wb: Workbook, ws_name: str, insurer_name: str, res: dict, pm: dict) -> None:
    wb.create_sheet(ws_name)
    ws = wb[ws_name]

    row = [1]  # mutable row pointer

    def _next():
        r = row[0]
        row[0] += 1
        return r

    def title(text: str, font=TITLE_FONT, fill=None):
        r = _next()
        ws.cell(r, 1, text).font = font
        if fill:
            ws.cell(r, 1).fill = fill

    def header(text: str):
        r = _next()
        ws.cell(r, 1, text).font = HEADER_FONT
        ws.cell(r, 1).fill = SECTION_FILL

    def data_row(label: str, value, comment: str = "", bold: bool = False, fill=None):
        r = _next()
        c = ws.cell(r, 1, label)
        if bold:
            c.font = Font(bold=True)
        if fill:
            c.fill = fill
        ws.cell(r, 2, value if value is not None else "")
        if comment:
            ws.cell(r, 3, comment)

    def blank(n: int = 1):
        row[0] += n

    def sv(d: dict | None, key: str):
        if d is None:
            return None
        v = d.get(key)
        return v  # None renders as blank cell

    s1  = res.get("s1")
    s2a = res.get("s2a")
    s2d = res.get("s2d")
    s3  = res.get("s3")
    s3d = res.get("s3d")
    s4  = res.get("s4")
    s5  = res.get("s5")
    s6  = res.get("s6")

    # Title rows
    r = _next()
    ws.cell(r, 1, f"Mapping Financial Statements — {insurer_name}").font = TITLE_FONT
    r = _next()
    ws.cell(r, 1, "SOTI  FY2025").font = Font(italic=True)
    r = _next()
    for col, val in enumerate([insurer_name, "FY2025", "Source"], start=1):
        ws.cell(r, col, val).font = Font(bold=True)

    blank()

    # ---- (1) PORTFOLIO OVERVIEW ------------------------------------------------
    header("(1) OVERVIEW OF PORTFOLIO")
    blank()
    header("(i) LIFE")
    data_row("General Measurement Model",   sv(s1, "life_gmm"),    "Note 7")
    data_row("Variable Fee Approach",       sv(s1, "life_vfa"),    "Note 7")
    data_row("Premium Allocation Approach", sv(s1, "life_paa"),    "Note 7")
    data_row("Subtotal Life",               sv(s1, "life_total"),  "Subtotal", bold=True, fill=SUBTOT_FILL)
    blank()
    header("(ii) NON-LIFE")
    data_row("General Measurement Model",   sv(s1, "nonlife_gmm"),    "Note 7")
    data_row("Premium Allocation Approach", sv(s1, "nonlife_paa"),    "Note 7")
    data_row("Subtotal Non-Life",           sv(s1, "nonlife_total"),  "Subtotal", bold=True, fill=SUBTOT_FILL)
    blank()
    header("(iii) HEALTH")
    data_row("General Measurement Model",   sv(s1, "health_gmm"),    "Note 7")
    data_row("Premium Allocation Approach", sv(s1, "health_paa"),    "Note 7")
    data_row("Subtotal Health",             sv(s1, "health_total"),  "Subtotal", bold=True, fill=SUBTOT_FILL)
    blank()
    header("(iv) TOTAL DIRECT")
    data_row("General Measurement Model",   sv(s1, "total_gmm"),    "Subtotal")
    data_row("Variable Fee Approach",       sv(s1, "total_vfa"),    "Subtotal")
    data_row("Premium Allocation Approach", sv(s1, "total_paa"),    "Subtotal")
    data_row("TOTAL",                       sv(s1, "total_direct"), "Subtotal", bold=True, fill=SUBTOT_FILL)
    blank()
    data_row("Total reinsurance ceded (net)", sv(s1, "reins_total"), "Note 7")
    blank(2)

    # ---- (2) FINANCIAL PERFORMANCE ---------------------------------------------
    header("(2) FINANCIAL PERFORMANCE")
    blank()
    data_row("a) Net insurance service result",   sv(s2a, "net_insurance_result"),  "P&L / Note 10")
    data_row("  1) CSM release",                  sv(s2a, "csm_release_total"),     "Note 10")
    data_row("  2) RA release",                   sv(s2a, "ra_release_total"),      "Note 10")
    data_row("  3) PAA insurance revenue",        sv(s2a, "paa_insurance_revenue"), "Note 10")
    data_row("  4) PAA incurred claims",          sv(s2a, "paa_incurred_claims"),   "Note 10")
    data_row("  5) GMM incurred claims",          sv(s2a, "gmm_incurred_claims"),   "Note 10")
    blank()
    data_row("b) Net reinsurance result",         sv(s2a, "net_reinsurance_result"), "P&L")
    data_row("c) Net financial result (ins act)", sv(s2a, "net_financial_result"),   "P&L")
    blank()
    data_row("PROFIT BEFORE TAX",                 sv(s2a, "profit_before_tax"),  "P&L", bold=True, fill=SUBTOT_FILL)
    blank()
    data_row("Insurance revenue (total)",         sv(s2a, "insurance_revenue_total"), "P&L")
    blank(2)

    # ---- (3) DISCOUNT RATES ----------------------------------------------------
    header("(3) OVERVIEW OF DISCOUNT RATES / CURVES")
    blank()
    header("a) Liquid curve (GMM)")
    for yr, key in [("1 year","liquid_1y"),("5 years","liquid_5y"),("10 years","liquid_10y"),
                    ("15 years","liquid_15y"),("20 years","liquid_20y"),("30 years","liquid_30y"),
                    ("40 years","liquid_40y"),("50 years","liquid_50y")]:
        data_row(yr, sv(s2d, key), "Discount note")
    blank()
    header("b) Illiquid curve (Life / VFA)")
    for yr, key in [("1 year","illiquid_1y"),("5 years","illiquid_5y"),("10 years","illiquid_10y"),
                    ("15 years","illiquid_15y"),("20 years","illiquid_20y"),("30 years","illiquid_30y"),
                    ("40 years","illiquid_40y"),("50 years","illiquid_50y")]:
        data_row(yr, sv(s2d, key), "Discount note")
    blank()
    header("c) Supplementary")
    data_row("Ultimate Forward Rate (UFR)",   sv(s2d, "ufr"),                  "Discount note")
    data_row("First Smoothing Point (years)", sv(s2d, "fsp_years"),            "Discount note")
    data_row("Cost of capital rate",          sv(s2d, "cost_of_capital_rate"), "RA section")
    data_row("Confidence — Non-Life",         sv(s2d, "confidence_nonlife"),   "RA section")
    data_row("Confidence — Life",             sv(s2d, "confidence_life"),      "RA section")
    data_row("Confidence — Health",           sv(s2d, "confidence_health"),    "RA section")
    blank(2)

    # ---- (4) CSM AND RA ROLLFORWARDS -------------------------------------------
    header("(4) LRC / LIC DEEP-DIVE")
    blank()
    header("(i.a) CSM DEVELOPMENT — TOTAL")
    data_row("a) Opening balance",          sv(s3, "csm_opening_total"),          "Note 7")
    data_row("b) New business",             sv(s3, "csm_new_business"),           "Note 7")
    data_row("c) Finance result",           sv(s3, "csm_finance_result"),         "Note 7")
    data_row("d) Future service changes",   sv(s3, "csm_future_service_changes"), "Note 7")
    data_row("e) CSM release",              sv(s3, "csm_release"),                "Note 7")
    data_row("f) Other / FX",              sv(s3, "csm_other"),                  "Note 7")
    data_row("g) Closing balance",          sv(s3, "csm_closing_total"),          "Note 7", bold=True, fill=SUBTOT_FILL)
    blank()
    header("(i.b) CSM — NON-LIFE SPLIT")
    data_row("Opening",      sv(s3, "csm_opening_nonlife"),     "Note 7")
    data_row("New business", sv(s3, "csm_new_business_nonlife"),"Note 7")
    data_row("Release",      sv(s3, "csm_release_nonlife"),     "Note 7")
    data_row("Closing",      sv(s3, "csm_closing_nonlife"),     "Note 7")
    blank()
    header("(i.c) CSM — LIFE SPLIT")
    data_row("Opening",      sv(s3, "csm_opening_life"),     "Note 7")
    data_row("New business", sv(s3, "csm_new_business_life"),"Note 7")
    data_row("Release",      sv(s3, "csm_release_life"),     "Note 7")
    data_row("Closing",      sv(s3, "csm_closing_life"),     "Note 7")
    blank()
    header("(i.d) CSM MATURITY")
    data_row("0-1 year",    sv(s3d, "csm_maturity_lt1y"),   "Note 7")
    data_row("1-5 years",   sv(s3d, "csm_maturity_1to5y"),  "Note 7")
    data_row("5-10 years",  sv(s3d, "csm_maturity_5to10y"), "Note 7")
    data_row(">10 years",   sv(s3d, "csm_maturity_gt10y"),  "Note 7")
    data_row("Total",       sv(s3d, "csm_maturity_total"),  "Subtotal", bold=True, fill=SUBTOT_FILL)
    blank(2)

    header("(ii.a) RISK ADJUSTMENT DEVELOPMENT — TOTAL")
    data_row("a) Opening balance",              sv(s3, "ra_opening_total"),          "Note 7")
    data_row("b) New business",                 sv(s3, "ra_new_business"),           "Note 7")
    data_row("c) Finance result",               sv(s3, "ra_finance_result"),         "Note 7")
    data_row("d.i) Future svc (adj CSM)",       sv(s3, "ra_future_service_adj_csm"), "Note 7")
    data_row("d.ii) Future svc (no CSM)",       sv(s3, "ra_future_service_no_csm"),  "Note 7")
    data_row("e) Past service",                 sv(s3, "ra_past_service"),           "Note 7")
    data_row("f) RA release",                   sv(s3, "ra_release"),                "Note 7")
    data_row("g) Other / FX",                  sv(s3, "ra_other"),                  "Note 7")
    data_row("h) Closing balance",              sv(s3, "ra_closing_total"),          "Note 7", bold=True, fill=SUBTOT_FILL)
    blank()
    data_row("RA Non-Life — opening", sv(s3, "ra_opening_nonlife"), "Note 7")
    data_row("RA Non-Life — closing", sv(s3, "ra_closing_nonlife"), "Note 7")
    data_row("RA Life — opening",     sv(s3, "ra_opening_life"),    "Note 7")
    data_row("RA Life — closing",     sv(s3, "ra_closing_life"),    "Note 7")
    blank(2)

    header("(iii) LOSS COMPONENT")
    data_row("Opening total",          sv(s3, "lc_opening_total"),          "Note 7")
    data_row("Losses recognised",      sv(s3, "lc_losses_recognised"),      "Note 7")
    data_row("Systematic allocation",  sv(s3, "lc_systematic_alloc"),       "Note 7")
    data_row("Future service changes", sv(s3, "lc_future_service_changes"), "Note 7")
    data_row("Finance",                sv(s3, "lc_finance"),                "Note 7")
    data_row("Other",                  sv(s3, "lc_other"),                  "Note 7")
    data_row("Closing total",          sv(s3, "lc_closing_total"),          "Note 7", bold=True, fill=SUBTOT_FILL)
    blank()
    data_row("Non-Life LC opening",    sv(s3, "lc_opening_nonlife"), "Note 7")
    data_row("Non-Life LC closing",    sv(s3, "lc_closing_nonlife"), "Note 7")
    data_row("Life LC opening",        sv(s3, "lc_opening_life"),    "Note 7")
    data_row("Life LC closing",        sv(s3, "lc_closing_life"),    "Note 7")
    data_row("Health LC opening",      sv(s3, "lc_opening_health"),  "Note 7")
    data_row("Health LC closing",      sv(s3, "lc_closing_health"),  "Note 7")
    blank(2)

    # ---- (5) SOLVENCY ----------------------------------------------------------
    header("(5) CAPITAL POSITIONS")
    blank()
    data_row("Solvency II ratio",           sv(s4, "solvency2_ratio"),    "SII section")
    data_row("Solvency II target ratio",    sv(s4, "solvency2_target"),   "SII section")
    data_row("SCR (EUR m)",                 sv(s4, "scr"),                "SII section")
    data_row("Eligible own funds (EUR m)",  sv(s4, "eligible_own_funds"), "SII section")
    data_row("Capital generated (EUR m)",   sv(s4, "capital_generated"),  "SII section")
    blank(2)

    # ---- (6) INVESTMENTS -------------------------------------------------------
    header("(6) INVESTMENT ASSET MIX")
    blank()
    data_row("Equities",           sv(s5, "equities"),          "Investments note")
    data_row("Government bonds",   sv(s5, "govt_bonds"),        "Investments note")
    data_row("Corporate bonds",    sv(s5, "corporate_bonds"),   "Investments note")
    data_row("Mortgages",          sv(s5, "mortgages"),         "Investments note")
    data_row("Other fixed income", sv(s5, "other_fixed_income"),"Investments note")
    data_row("Derivatives (net)",  sv(s5, "derivatives_net"),   "Investments note")
    data_row("Other",              sv(s5, "other_investments"), "Investments note")
    data_row("TOTAL",              sv(s5, "total_investments"), "Subtotal", bold=True, fill=SUBTOT_FILL)
    blank()
    data_row("FVOCI",          sv(s5, "fvoci"),         "Investments note")
    data_row("FVTPL",          sv(s5, "fvtpl"),         "Investments note")
    data_row("Amortised cost", sv(s5, "amortised_cost"),"Investments note")
    blank(2)

    # ---- (7) PREMIUM -----------------------------------------------------------
    header("(7) GROSS WRITTEN PREMIUM (NON-GAAP)")
    blank()
    data_row("Life",          sv(s6, "gwp_life"),    "Results section")
    data_row("Health",        sv(s6, "gwp_health"),  "Results section")
    data_row("Non-Life",      sv(s6, "gwp_nonlife"), "Results section")
    data_row("Pensions",      sv(s6, "gwp_pensions"),"Results section")
    data_row("International", sv(s6, "gwp_intl"),    "Results section")
    data_row("TOTAL",         sv(s6, "gwp_total"),   "Subtotal", bold=True, fill=SUBTOT_FILL)

    # Column formatting
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 35
    ws.freeze_panes = "B4"

    # Right-align the value column
    for r in range(1, row[0]):
        ws.cell(r, 2).alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("\n" + "=" * 60)
    print("IFRS 17 Extraction — All Insurers 2025")
    print("=" * 60)
    print(f"Extracting {len(PAGE_MAPS)} insurers in parallel (ThreadPoolExecutor)...")

    t0 = time.time()

    # ThreadPoolExecutor is safe here: threads share memory, no fork issues.
    # One worker per insurer gives ~4x speedup vs sequential.
    results = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(PAGE_MAPS)) as executor:
        futures = {
            executor.submit(extract_insurer, pm): key
            for key, pm in PAGE_MAPS.items()
        }
        for future in concurrent.futures.as_completed(futures):
            key = futures[future]
            try:
                results[key] = future.result()
            except Exception as e:
                print(f"WARNING: {PAGE_MAPS[key]['short_name']} extraction failed: {e}")
                results[key] = None

    elapsed = round(time.time() - t0)
    print(f"\nAll extractions done in {elapsed}s.")

    wb = Workbook()
    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for key, pm in PAGE_MAPS.items():
        ws_name = pm["short_name"]
        res = results.get(key)
        if res is None:
            print(f"WARNING: {ws_name} extraction failed — skipping tab.")
            continue
        print(f"Assembling tab: {ws_name}")
        write_insurer_tab(wb, ws_name, ws_name, res, pm)

    print("\nSaving workbook...")
    wb.save(OUT_PATH)
    print(f"Done. Output written to: {OUT_PATH}")


if __name__ == "__main__":
    main()
